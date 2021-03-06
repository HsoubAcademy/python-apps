import openpyxl, smtplib
from pathlib import Path

# connect with Excelfile
file = openpyxl.load_workbook(Path.home() / Path('Desktop', 'members.xlsx'))
sheets = file.sheetnames
sheet = file['Sheet1']
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value
print(latestMonth)

# find unpaid members
unpaidMembers = {}

for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Send Email
smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.starttls()
sender_email = "academyhsoub1@gmail.com"
password = input(str("Please enter your password: "))
smtpObj.login(sender_email, password)

for name, email in unpaidMembers.items():
    body = """Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues 
    for %s.Please make this payment as soon as possible.Thank you!'""" %(latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail(sender_email, email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

smtpObj.quit()